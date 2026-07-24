# -*- coding: utf-8 -*-
"""
Projecao de dias de estoque por insumo (loja Domino's).

Entradas (layout KEYS, so mudam loja e data):
  - KEYS Inventory Item Trend : historico diario -> coluna 'Ideal Usage' (consumo por ficha tecnica)
  - KEYS Inventory Variance   : foto do inventario -> coluna 'Ending Inventory Qty' (estoque inicial/final)

Logica:
  1. Media de 'Ideal Usage' por DIA DA SEMANA (seg..dom), por insumo, a partir do historico.
  2. Estoque inicial = 'Ending Inventory Qty' do Variance, na data do arquivo Variance.
  3. Simula o consumo dia a dia a partir da data do Variance, descontando a media do
     respectivo dia da semana, ate o estoque zerar.
  Saida: dias de estoque e a data/dia da semana em que o estoque acabaria.

Premissa: o consumo comeca a contar NO dia do arquivo Variance (data do arquivo = dia 0),
pois 'Ending Inventory Qty' e o estoque inicial dessa data.
"""
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import glob, os, sys

PASTA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "são luis")
DIAS_PT = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
MAX_DIAS = 400  # teto de simulacao

# Massas exibidas em BANDEJAS (Inv Code -> unidades por bandeja).
# So afeta a exibicao (estoque e medias); dias e datas nao mudam.
BANDEJA = {
    "7DOUGH": 12,    # MASSA 7
    "8HDOUGH": 12,   # MASSA 8.5
    "11HDOUGH": 8,   # MASSA 11.5
    "14DOUGH": 6,    # MASSA 14
    "11HDOPAN": 7,   # MASSA PAN 11.5
}


def carregar(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = rows[0]
    idx = {h: i for i, h in enumerate(hdr)}
    data = [r for r in rows[1:] if r[0] is not None]
    return idx, data


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def achar(padrao):
    m = glob.glob(os.path.join(PASTA, padrao))
    if not m:
        sys.exit("Arquivo nao encontrado: " + padrao)
    return m[0]


def main():
    f_trend = achar("*Item Trend*.xlsx")
    f_var = achar("*Variance*.xlsx")

    # ---- 1. Media de Ideal Usage por dia da semana, por insumo ----
    idx, data = carregar(f_trend)
    somas = defaultdict(lambda: [0.0] * 7)   # inv_code -> soma por weekday
    contagem = defaultdict(lambda: [0] * 7)  # inv_code -> n dias por weekday
    desc = {}
    unid = {}
    for r in data:
        code = r[idx["Inv Code"]]
        d = r[idx["Date"]]
        if not isinstance(d, datetime):
            d = datetime.strptime(str(d)[:10], "%Y-%m-%d")
        wd = d.weekday()
        somas[code][wd] += num(r[idx["Ideal Usage"]])
        contagem[code][wd] += 1
        desc.setdefault(code, r[idx["Description"]])
        unid.setdefault(code, r[idx["Count Unit"]])

    media_wd = {}
    for code in somas:
        media_wd[code] = [
            (somas[code][w] / contagem[code][w]) if contagem[code][w] else 0.0
            for w in range(7)
        ]

    # ---- 2. Estoque inicial (Ending Inventory Qty) e data do Variance ----
    idxv, datav = carregar(f_var)
    # data do arquivo Variance: extrai do nome do arquivo (aaaa-mm-dd)
    import re
    base = os.path.basename(f_var)
    m = re.search(r"(\d{4}-\d{2}-\d{2})", base)
    data_arquivo = datetime.strptime(m.group(1), "%Y-%m-%d")
    # 'Ending Inventory Qty' e o estoque FINAL da data do arquivo = estoque INICIAL
    # do dia seguinte. Logo o consumo comeca a contar no dia seguinte.
    data_base = data_arquivo + timedelta(days=1)
    ml = re.search(r"Store-(\d+)", base)
    loja = ml.group(1) if ml else "loja"
    estoque = {}
    for r in datav:
        code = r[idxv["Inventory Code"]]
        estoque[code] = num(r[idxv["Ending Inventory Qty"]])

    # ---- 3. Simulacao do consumo dia a dia ----
    resultados = []
    for code, est in estoque.items():
        med = media_wd.get(code)
        d = desc.get(code, "")
        u = unid.get(code, "")
        if med is None:
            resultados.append((code, d, u, est, None, None, None, "sem histórico"))
            continue
        cons_semana = sum(med)
        if cons_semana <= 0:
            resultados.append((code, d, u, est, 0.0, None, None, "sem consumo"))
            continue

        stock = est
        k = 0
        dias_total = None
        data_fim = None
        while k < MAX_DIAS:
            wd = (data_base + timedelta(days=k)).weekday()
            uso = med[wd]
            if uso <= 0:
                k += 1
                continue
            if stock >= uso:
                stock -= uso
                k += 1
            else:
                frac = stock / uso
                dias_total = k + frac
                data_fim = data_base + timedelta(days=k)
                break
        if dias_total is None:
            resultados.append((code, d, u, est, float(MAX_DIAS),
                               None, None, ">%d dias" % MAX_DIAS))
        else:
            resultados.append((code, d, u, est, round(dias_total, 1),
                               data_fim, DIAS_PT[data_fim.weekday()], ""))

    # ordena por dias de estoque (mais urgente primeiro)
    resultados.sort(key=lambda x: (x[4] is None, x[4] if x[4] is not None else 0))

    # ---- Saida Excel ----
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projeção de Estoque"
    cab = ["Inv Code", "Descrição", "Unid.", "Estoque atual",
           "Dias de estoque", "Data em que acaba", "Dia da semana", "Obs.",
           "Média Seg", "Média Ter", "Média Qua", "Média Qui",
           "Média Sex", "Média Sáb", "Média Dom"]
    ws.append(cab)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for (code, d, u, est, dias, dfim, dsem, obs) in resultados:
        med = media_wd.get(code, [None] * 7)
        fator = BANDEJA.get(code)
        if fator:
            u = "bandeja"
            est_ex = est / fator
            med_ex = [(x / fator if x is not None else None) for x in med]
        else:
            est_ex = est
            med_ex = med
        ws.append([code, d, u, round(est_ex, 2), dias,
                   dfim.strftime("%d/%m/%Y") if dfim else "", dsem, obs]
                  + [round(x, 2) if x is not None else "" for x in med_ex])
    # cor de alerta por dias de estoque
    verm = PatternFill("solid", fgColor="F8CBAD")
    amar = PatternFill("solid", fgColor="FFE699")
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        v = row[0].value
        if isinstance(v, (int, float)):
            if v <= 2:
                row[0].fill = verm
            elif v <= 5:
                row[0].fill = amar
    larguras = [11, 26, 6, 13, 14, 16, 13, 14] + [9] * 7
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    out = os.path.join(os.path.dirname(f_var),
                       "Projeção Estoque %s (%s).xlsx" %
                       (loja or "loja", data_arquivo.strftime("%Y-%m-%d")))
    wb.save(out)

    # resumo no console
    print("Data do arquivo (Variance):", data_arquivo.strftime("%d/%m/%Y"),
          "-", DIAS_PT[data_arquivo.weekday()], "(estoque final)")
    print("Consumo comeca em:", data_base.strftime("%d/%m/%Y"),
          "-", DIAS_PT[data_base.weekday()], "(estoque inicial)")
    print("Insumos projetados:", len(resultados))
    print("Arquivo gerado:", out)
    print()
    print("%-9s %-24s %8s %6s  %-12s %s" %
          ("Code", "Descrição", "Estoque", "Dias", "Acaba em", "Dia"))
    for (code, d, u, est, dias, dfim, dsem, obs) in resultados[:20]:
        fator = BANDEJA.get(code)
        est_ex = est / fator if fator else est
        unid = " band" if fator else ""
        print("%-9s %-24s %8.2f%-5s %6s  %-12s %s" %
              (code, str(d)[:24], est_ex, unid,
               ("%.1f" % dias) if dias is not None else "-",
               dfim.strftime("%d/%m/%Y") if dfim else (obs or "-"),
               dsem or ""))


if __name__ == "__main__":
    main()
